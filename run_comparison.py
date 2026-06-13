"""
run_comparison.py
-----------------
Run Windpower Lite for all 10 turbines across 5 Nigerian locations
and save results to both CSV and Excel.

Run from the project root:
    python run_comparison.py

Edit the CONFIGURATION section to change sites, date range, or output paths.
"""

import os
import sys
import datetime

import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))

from main import run_all_turbines

# ── CONFIGURATION ─────────────────────────────────────────────────────────────
LOCATIONS = [
    { "name": "Sokoto — Excellent Wind",   "lat": 13.0059, "lon": 5.2476  },
    { "name": "Katsina — Good Wind",       "lat": 12.9908, "lon": 7.6018  },
    { "name": "Ilorin — Medium Wind",      "lat":  8.4966, "lon": 4.5421  },
    { "name": "Abuja — Poor Wind",         "lat":  9.0765, "lon": 7.3986  },
    { "name": "Port Harcourt — Very Poor", "lat":  4.8156, "lon": 7.0498  },
]

START_DATE  = datetime.datetime(2023, 1,  1)
END_DATE    = datetime.datetime(2023, 12, 31)
USE_HYBRID  = True

OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "results")
CSV_FILE    = os.path.join(OUTPUT_DIR, "windpower_lite_all_turbines.csv")
EXCEL_FILE  = os.path.join(OUTPUT_DIR, "windpower_lite_all_turbines.xlsx")


# ── FORMATTING ────────────────────────────────────────────────────────────────
def format_results(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {
        "location":                    "Location",
        "turbine":                     "Turbine Model",
        "rated_power_kw":              "Rated Power (kW)",
        "rotor_diameter_m":            "Rotor Diameter (m)",
        "hub_height_m":                "Hub Height (m)",
        "aep_physics_kwh":             "AEP — Physics (kWh)",
        "capacity_factor_physics_pct": "CF — Physics (%)",
        "aep_hybrid_kwh":              "AEP — Hybrid (kWh)",
        "capacity_factor_hybrid_pct":  "CF — Hybrid (%)",
    }
    df = df.rename(columns=rename_map)

    if "AEP — Physics (kWh)" in df.columns and "AEP — Hybrid (kWh)" in df.columns:
        df["Hybrid Uplift (%)"] = (
            (df["AEP — Hybrid (kWh)"] - df["AEP — Physics (kWh)"])
            / df["AEP — Physics (kWh)"]
            * 100
        ).round(1)

    return df


def save_excel(df: pd.DataFrame, path: str) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  Excel skipped — install openpyxl: pip install openpyxl")
        return

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="All Locations")
        ws = writer.sheets["All Locations"]

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = max(
                max_len + 4, 14
            )

        # Header styling
        header_fill = PatternFill("solid", fgColor="0A7E8C")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font  = header_fill and header_font
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Alternate row shading
        light_fill = PatternFill("solid", fgColor="E2EAF0")
        for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row_cells:
                    cell.fill = light_fill

    print(f"  Excel saved  →  {path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 70)
    print("  WINDPOWER LITE — 5-LOCATION COMPARATIVE RUN")
    print("=" * 70)
    print(f"  Period  : {START_DATE.date()} → {END_DATE.date()}")
    print(f"  Hybrid  : {'Yes' if USE_HYBRID else 'No'}")
    print(f"  Locations: {len(LOCATIONS)}")
    print(f"  Turbines : 10 Vestas configurations")
    print("=" * 70 + "\n")

    all_results = []

    for loc in LOCATIONS:
        print(f"\n{'─'*70}")
        print(f"  Running: {loc['name']}")
        print(f"  Coords : {loc['lat']}°N, {loc['lon']}°E")
        print(f"{'─'*70}")

        df = run_all_turbines(
            lat=loc["lat"],
            lon=loc["lon"],
            start_dt=START_DATE,
            end_dt=END_DATE,
            apply_hybrid_if_available=USE_HYBRID,
        )
        df.insert(0, "location", loc["name"])
        all_results.append(df)

    # Combine all locations
    df_combined = pd.concat(all_results, ignore_index=True)
    df_formatted = format_results(df_combined)

    # ── Print summary table ───────────────────────────────────────────────────
    print("\n\n" + "=" * 70)
    print("  FULL COMPARATIVE RESULTS — ALL LOCATIONS AND TURBINES")
    print("=" * 70)
    print(df_formatted.to_string(index=False))

    # ── Quick stats ───────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("  QUICK STATS")
    print("=" * 70)

    hybrid_col = "AEP — Hybrid (kWh)"
    cf_col     = "CF — Hybrid (%)"

    if hybrid_col in df_formatted.columns:
        best  = df_formatted.loc[df_formatted[hybrid_col].idxmax()]
        worst = df_formatted.loc[df_formatted[hybrid_col].idxmin()]
        print(f"\n  Highest AEP : {best['Location']} — {best['Turbine Model']}")
        print(f"    AEP: {best[hybrid_col]:>12,.0f} kWh  CF: {best[cf_col]:.2f}%")
        print(f"\n  Lowest  AEP : {worst['Location']} — {worst['Turbine Model']}")
        print(f"    AEP: {worst[hybrid_col]:>12,.0f} kWh  CF: {worst[cf_col]:.2f}%")

    if "Hybrid Uplift (%)" in df_formatted.columns:
        avg_uplift = df_formatted["Hybrid Uplift (%)"].mean()
        print(f"\n  Average hybrid uplift : {avg_uplift:.1f}%")

    # ── Save outputs ──────────────────────────────────────────────────────────
    print("\nSaving outputs ...")
    df_formatted.to_csv(CSV_FILE, index=False)
    print(f"  CSV saved    →  {CSV_FILE}")
    save_excel(df_formatted, EXCEL_FILE)

    print("\n" + "=" * 70)
    print("  Done. Results ready for Chapter 4.")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()